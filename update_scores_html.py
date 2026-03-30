import traceback
import pandas as pd
import app
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
import logging
from datetime import datetime
from openpyxl.styles import PatternFill, Font  
from openpyxl.utils import get_column_letter  
import shutil  
import os  
import sqlite3
from datetime import datetime
import pytz


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)


def insert_log_message(message):
    try:
        #logging.info(f"Inserting logs")
        # Create SQLite connection
        #conn_logs = sqlite3.connect('/mnt/sqlite/cricbattle.db' if os.environ.get("WEBSITE_SITE_NAME") else 'instance/cricbattle.db')
        conn_logs = sqlite3.connect('/mnt/sqlite/cricbattle.db' if os.environ.get("WEBSITE_SITE_NAME") else '/mnt/sqlite/cricbattle.db' if os.environ.get("GOOGLE_CLOUD_PROJECT") else 'instance/cricbattle.db')  

        # Create a cursor object 
        cursor = conn_logs.cursor()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Convert DataFrame to string if message is a DataFrame
        if isinstance(message, pd.DataFrame):
            message = message.to_string()
            
        # Insert the log message into the logs table
        cursor.execute("INSERT INTO logs (timestamp, message) VALUES (?, ?)", (timestamp, str(message)))
        # Commit the changes
        conn_logs.commit()
        # Close the connection
        conn_logs.close()
        #logging.info(f"Inserted logs")
    except Exception as e:
        logging.error(f"Error inserting logs: {str(e)}") 


def adjust_column_widths(sheet):  
    logging.info("Adjusting column widths")  
    for column in sheet.columns:  
        max_length = 0  
        column = [cell for cell in column]  
        for cell in column:  
            try:  
                if len(str(cell.value)) > max_length:  
                    max_length = len(str(cell.value))  
            except:  
                pass  
        adjusted_width = (max_length + 2)  
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width  

  


def read_excel_file(filename):
    try:
        return pd.read_excel(filename)
    except FileNotFoundError:
        print(f"Error: The file '{filename}' was not found.")
    except Exception as e:
        print(f"An error occurred while reading '{filename}': {e}")
    return None

def calculate_best_11_old(df):  
    #logging.info("Calculating best 11 players for each team")  
    best_11 = []  
    #print(df)
      
    for team, group in df.groupby('Team Name'):  
        players = group.copy()  
        players.sort_values(by='TotalScore', ascending=False, inplace=True)  
        #logging.info(f"Calculating best 11 players for team {team}") 
          
        # Initialize constraints  
        wk_needed = 1  # At least 1 WK required
        bat_needed = 4  # At least 4 batters, including WK  
        all_needed = 1  
        bowl_needed = 3  
        max_overseas = 4
        max_ipl_team = 3  
          
        # Create best 11 list  
        selected = []  
        selected_ids = set()  # To track selected player IDs or names  
        overseas_counter = 0  # Counter for overseas players
        ipl_team_counts = {}  # Track number of players from each IPL team
  
        # Role-specific selections with adjusted logic  
        wk_count = 0  # Track selected WK count  
        bat_count = 0  # Track selected batters (including WKs)  
        
        for _, player in players.iterrows():  
            #logging.info(f"Processing player {player['Player Name']}")
            if len(selected) >= 11:  
                break  
            
            player_role = player['Role']  
            player_is_overseas = player['foreign_player']  
            player_ipl_team = player['IPL Team']
            player_id = player['PlayerId']  
            
            if player_id in selected_ids or overseas_counter >= max_overseas or ipl_team_counts.get(player_ipl_team, 0) >= max_ipl_team:
                continue
            
            if player_role == 'Wicket-Keeper':  
                if wk_count < 1:  # Allow multiple WKs up to 4  
                    selected.append(player)  
                    selected_ids.add(player_id)  
                    wk_count += 1  
                    bat_count += 1  # WK is also counted as a batter  
                    ipl_team_counts[player_ipl_team] = ipl_team_counts.get(player_ipl_team, 0) + 1  
                    if player_is_overseas:
                        overseas_counter = overseas_counter + 1
            
            elif player_role == 'All-Rounder' and all_needed > 0:  
                selected.append(player)  
                selected_ids.add(player_id)  
                all_needed -= 1   
                ipl_team_counts[player_ipl_team] = ipl_team_counts.get(player_ipl_team, 0) + 1  
                if player_is_overseas:
                        overseas_counter = overseas_counter + 1
                
            elif player_role == 'Bowler' and bowl_needed > 0:  
                selected.append(player)  
                selected_ids.add(player_id)  
                bowl_needed -= 1  
                ipl_team_counts[player_ipl_team] = ipl_team_counts.get(player_ipl_team, 0) + 1  
                if player_is_overseas:
                        overseas_counter = overseas_counter + 1
  
        #logging.info(f"IPL Team count so far {ipl_team_counts} for team {player_ipl_team}") 
        #logging.info(f"Overseas count so far {overseas_counter} for team {player_ipl_team}")
  
        # Ensure at least 1 WK is selected  
        for _, player in players.iterrows():  
            player_is_overseas = player['foreign_player']  
            player_ipl_team = player['IPL Team']
            if len(selected) >= 11:  
                break 
            if player['Role'] in ['Batsman', 'Wicket-Keeper'] and player['PlayerId'] not in selected_ids and len(selected) < 11:  
                selected.append(player)  
                selected_ids.add(player['PlayerId'])  
                wk_count += 1  
                bat_count += 1  
                ipl_team_counts[player_ipl_team] = ipl_team_counts.get(player_ipl_team, 0) + 1  
                if player_is_overseas:
                        overseas_counter = overseas_counter + 1
                #logging.info(f"Adding player {player['Player Name']} for team {team}")
                if bat_count >= bat_needed:  
                    break  
  

  
        # Ensure we have exactly 11 players by filling any gaps with highest scorers  
        while len(selected) < 11:  
            for _, player in players.iterrows():  
                player_is_overseas = player['foreign_player']  
                player_ipl_team = player['IPL Team']
                if len(selected) >= 11:  
                    break 
                player_id = player['PlayerId']  
                #logging.info(f"Re-Processing player {player['Player Name']}")
                if player_id in selected_ids or  (player_is_overseas and overseas_counter >= max_overseas) or ipl_team_counts.get(player_ipl_team, 0) >= max_ipl_team:
                    #logging.info(f"Player ID: {player_id}, Selected IDs: {selected_ids}, Overseas Counter: {overseas_counter}, Max Overseas: {max_overseas}, IPL Team: {player_ipl_team}, IPL Team Count: {ipl_team_counts.get(player_ipl_team, 0)}, Max IPL Team: {max_ipl_team}, Is Overseas: {player_is_overseas}") 
                    #logging.info(f"Skipping player {player['Player Name']} due to foreign player or IPL team logic")
                    continue  
                #logging.info(f"Adding player {player['Player Name']} to best 11") 
                selected.append(player)  
                selected_ids.add(player_id)  
                ipl_team_counts[player_ipl_team] = ipl_team_counts.get(player_ipl_team, 0) + 1  
                if player_is_overseas:
                        overseas_counter = overseas_counter + 1
  
        #logging.info(f"Finished best 11 players for team {team} - {len(selected)}")  
  
        # Sum the total score of best 11  
        best_11_points = sum(player['TotalScore'] for player in selected)  
        best_11.append((team, best_11_points, selected))  

    
    #exit()  
    return best_11

import random
from collections import defaultdict

def calculate_best_11(df, league="FPL"):  
    best_11 = []  

    if not league:
        league = "FPL"   

    #print(df)

    try:
        mapping = init_player_id_mapping()
    except Exception:
        mapping = []

    for liga, old_player, new_player in mapping:
        #print(liga, league, old_player, new_player)
        if liga == league: 
            #print(df[df['Team Name'] == "KR"])
            df = df[df['Player Name'] != old_player]
            df = df[df['Player Name'] != new_player]
            #print("Removing old and new player from dataframe")

            #print(df[df['Team Name'] == "KR"])

    for team, group in df.groupby('Team Name'):  
        players = group.copy()  
        players.sort_values(by='TotalScore', ascending=False, inplace=True)  

        # Constraints  
        min_wk = 1
        min_bat = 4
        min_all = 1
        min_bowl = 3
        max_overseas = 4
        max_ipl_team = 3  

        # Selected Players  
        selected = []  
        selected_ids = set()  
        overseas_counter = 0  
        ipl_team_counts = defaultdict(int)  
        
        # Role Counts  
        wk_count, bat_count, all_count, bowl_count = 0, 0, 0, 0  

        def can_select(player):
            """Checks whether a player can be selected based on constraints."""
            if player['PlayerId'] in selected_ids:
                return False
            if overseas_counter >= max_overseas and player['foreign_player']:
                return False
            if ipl_team_counts[player['IPL Team']] >= max_ipl_team:
                return False
            return True  

        def add_player(player):
            """Adds a player to the best 11 if valid."""
            nonlocal overseas_counter, wk_count, bat_count, all_count, bowl_count  

            if len(selected) >= 11:
                return False  

            if not can_select(player):
                return False  

            selected.append(player)  
            selected_ids.add(player['PlayerId'])  
            ipl_team_counts[player['IPL Team']] += 1  

            if player['foreign_player']:
                overseas_counter += 1  

            # Role tracking  
            if player['Role'] == 'Wicket-Keeper':
                wk_count += 1
                bat_count += 1  # WK counts as a batter
            elif player['Role'] == 'Batsman':
                bat_count += 1
            elif player['Role'] == 'All-Rounder':
                all_count += 1
            elif player['Role'] == 'Bowler':
                bowl_count += 1  

            return True  

        # **Step 1: Select Players Based on Role Requirements**  
        for _, player in players.iterrows():  
            if player['Role'] == 'Wicket-Keeper' and wk_count < min_wk:
                add_player(player)  
            elif player['Role'] == 'All-Rounder' and all_count < min_all:
                add_player(player)  
            elif player['Role'] == 'Bowler' and bowl_count < min_bowl:
                add_player(player)  

        # **Step 2: Ensure At Least 4 Batters**
        for _, player in players.iterrows():
            if len(selected) >= 11:
                break
            if player['Role'] in ['Batsman', 'Wicket-Keeper'] and bat_count < min_bat:
                add_player(player)  

        # **Step 3: Fill Remaining Spots with Highest Available Players**
        for _, player in players.iterrows():
            if len(selected) >= 11:
                break
            add_player(player)  

        # **Step 4: Post-selection optimisation — swap out 0-pt blockers for excluded high-scorers**
        # If a player with actual points couldn't be selected only because their IPL team slot
        # was already full (occupied by a 0-pt player), swap the dead weight out.
        unselected = players[~players['PlayerId'].isin(selected_ids)]
        for _, candidate in unselected.sort_values('TotalScore', ascending=False).iterrows():
            if candidate['TotalScore'] <= 0:
                break  # nothing useful left to try
            # Only interested in candidates blocked by the IPL-team cap
            if ipl_team_counts[candidate['IPL Team']] < max_ipl_team:
                continue  # wasn't the team cap that blocked them; skip
            # Look for a 0-pt selected player from the same IPL team to evict
            for i, sel in enumerate(selected):
                if sel['IPL Team'] != candidate['IPL Team']:
                    continue
                if sel['TotalScore'] > 0:
                    continue  # don't evict a scorer
                # Simulate the swap and check all minimums are still satisfied
                new_wk   = wk_count   - (1 if sel['Role'] == 'Wicket-Keeper' else 0) \
                                       + (1 if candidate['Role'] == 'Wicket-Keeper' else 0)
                new_bat  = bat_count  - (1 if sel['Role'] in ['Batsman', 'Wicket-Keeper'] else 0) \
                                       + (1 if candidate['Role'] in ['Batsman', 'Wicket-Keeper'] else 0)
                new_all  = all_count  - (1 if sel['Role'] == 'All-Rounder' else 0) \
                                       + (1 if candidate['Role'] == 'All-Rounder' else 0)
                new_bowl = bowl_count - (1 if sel['Role'] == 'Bowler' else 0) \
                                       + (1 if candidate['Role'] == 'Bowler' else 0)
                # Overseas check
                new_overseas = overseas_counter \
                               - (1 if sel['foreign_player'] else 0) \
                               + (1 if candidate['foreign_player'] else 0)
                if (new_wk >= min_wk and new_bat >= min_bat and
                        new_all >= min_all and new_bowl >= min_bowl and
                        new_overseas <= max_overseas):
                    # Perform swap
                    selected[i] = candidate
                    selected_ids.discard(sel['PlayerId'])
                    selected_ids.add(candidate['PlayerId'])
                    wk_count, bat_count, all_count, bowl_count = new_wk, new_bat, new_all, new_bowl
                    overseas_counter = new_overseas
                    break  # move on to next unselected candidate

        # Calculate best 11 total score  
        best_11_points = sum(player['TotalScore'] for player in selected)  
        best_11.append((team, best_11_points, selected))  

    return best_11


def create_team_points_chart(team_points_df):
    fig = px.bar(
        team_points_df,
        x='Team Name',
        y=['TotalPoints', 'Best11Points'],
        title='Team Performance Comparison',
        barmode='group',
        labels={'value': 'Points', 'variable': 'Category'},
        color_discrete_sequence=['#1f77b4', '#ff7f0e']
    )
    return fig.to_html(full_html=False)

def create_player_performance_chart(player_team_points_df):
    top_players = player_team_points_df.nlargest(10, 'PlayerPoints')
    fig = px.bar(
        top_players,
        x='Player Name', 
        y='PlayerPoints',
        color='Team Name',
        title='Top 10 MVPs',
        labels={'PlayerPoints': 'Points'},
        text='PlayerPoints',
        text_auto=True
    )
    #fig.update_traces(textposition='outside')  
    return fig.to_html(full_html=False)

def create_role_distribution_chart(player_team_points_df):
    role_points = player_team_points_df.groupby('Role')['PlayerPoints'].sum().reset_index()
    fig = px.pie(
        role_points,
        values='PlayerPoints',
        names='Role',
        title='Points Distribution by Role',
        hole=0.3
    )
    return fig.to_html(full_html=False)

# Add row styling based on best 11 membership
def style_row(row, best_11_set):
    if (row['Team Name'], row['Player Name']) in best_11_set:
        return 'background-color: #e6ffe6'  # Light green background
    return ''

def generate_html_report(team_points_df, player_team_points_df, series_stats_df, scoreboard_stats_df, best_11_df, player_of_the_day, team_of_the_day, league, live_players_list, all_team_points_df, live_player_scores_df):
    
    team_chart = create_team_points_chart(team_points_df)
    player_chart = create_player_performance_chart(player_team_points_df)
    role_chart = create_role_distribution_chart(player_team_points_df)
    race_to_finish_chart = create_race_to_finish_chart(all_team_points_df)

    # Pre-compute ticker content (list of live players scrolling across top)
    if not live_players_list.empty:
        ticker_content = ''.join([
            f'<div class="ticker-item"><span class="player-name">{row["name"]}</span> <span class="score">{row["fpl_team"]}</span></div>'
            for _ in range(2) for _, row in live_players_list.iterrows()
        ])
    else:
        ticker_content = ''

    # Pre-compute conditional heading for live scores section
    scores_live_heading = '' if league == "JAL" else '<h2>Scores Live Today</h2>'

    # Create clickable player names with URLs and add background color for best 11 players
    player_team_points_df = player_team_points_df.sort_values(['Team Name', 'PlayerPoints'])
    
    # Create a set of (team, player) tuples from best_11_df for faster lookup
    best_11_set = set(zip(best_11_df['Team Name'], best_11_df['Player Name']))

    styled_df = player_team_points_df.copy()
    styled_df['Playing11'] = styled_df.apply(lambda x: 'Yes' if (x['Team Name'], x['Player Name']) in best_11_set else '', axis=1)    

    # Convert to HTML with styling
    styled_df['Player Name'] = styled_df.apply(
        lambda x: f'<a href="{x.PlayerId}" target="_blank">{x["Player Name"]}</a>', axis=1)
    
    # Drop PlayerId column
    styled_df = styled_df.drop('PlayerId', axis=1)


    # Convert DataFrames to HTML tables with Bootstrap styling

    team_table = team_points_df.to_html(classes=['table', 'table-striped', 'table-hover'], 
                                      index=False, 
                                      float_format=lambda x: '{:.2f}'.format(x) if pd.notnull(x) else '')
    player_table = styled_df.to_html(classes=['table', 'table-striped', 'table-hover'],
                                   index=False,
                                   float_format=lambda x: '{:.2f}'.format(x) if pd.notnull(x) else '',
                                   escape=False)    

    daily_scores_table = live_player_scores_df.to_html(classes=['table', 'table-striped', 'table-hover'],
                                   index=False,
                                   float_format=lambda x: '{:.2f}'.format(x) if pd.notnull(x) else '',
                                   escape=False)    

    # Convert series stats DataFrames to HTML tables
    series_tables = ""
    for key, df in series_stats_df.items():
        if key == "MOST_RUNS":
            key = "Best Batter"
        elif key == "MOST_WICKETS":
            key = "Best Bowler"
        elif key == "MOST_SIXES":
            key = "Most Sixes"
        else:
            logging.error(f"Unknown key {key}")

        #print(df)

        if df.empty:
            logging.error(f"Empty DataFrame for {key}")
            continue

        df['Team Name'] = df['Team Name'].fillna(value="LORDX1")
        del df['Player Name']
        cols = df.columns.tolist()
        #print(cols)
        cols = cols[1:2] + cols[:1] + cols[2:]        
        #print(cols)
        df = df[cols]
        series_tables += f"""
            <h2>{key}</h2>
            <div class="table-container">
                {df.head().to_html(classes=['table', 'table-striped', 'table-hover'], 
                           index=False,
                           float_format=lambda x: '{:.2f}'.format(x) if pd.notnull(x) else '')}
            </div>
        """

    # Convert series stats DataFrames to HTML tables
    sb_tables = ""
    for key, df in scoreboard_stats_df.items():
        
        if key == "Bat":
            key = "Most 50s Per Team"
        elif key == "Bowl":
            key = "Most 3fers Per Team"
        elif key == "Field":
            key = "Best Fielder"
        elif key == "POTM":
            key = "Most POTMs"
        else:
            logging.error(f"Unknown key {key}")

        #del df['Player Name']
        cols = df.columns.tolist()
        print(cols)
        sb_tables += f"""
            <h2>{key}</h2>
            <div class="table-container">
                {df.to_html(classes=['table', 'table-striped', 'table-hover'], 
                           index=True,
                           float_format=lambda x: '{:.2f}'.format(x) if pd.notnull(x) else '')}
            </div>
        """
    
   

    timestamp = datetime.now(pytz.timezone('Europe/Paris')).strftime("%Y-%m-%d %H:%M:%S %Z")     

    # extract player of the day and team of the day info for today
    """     'today': {'team': today_best_team[0], 'score': today_best_team[1]},
        'yesterday': {'team': yesterday_best_team[0], 'score': yesterday_best_team[1]}
    }
    and player of the day has
    return {
    'today': {
        'name': today_player_details.name if today_player_details else None,
        'team': today_player_details.team_name if today_player_details else None,
        'points': today_player.TotalScore if today_player else 0
    },
    'yesterday': {
        'name': yesterday_player_details.name if yesterday_player_details else None, 
        'team': yesterday_player_details.team_name if yesterday_player_details else None,
        'points': yesterday_player.TotalScore if yesterday_player else 0
    } """

    #print(player_of_the_day['today']['name'], player_of_the_day['today']['team'], player_of_the_day['today']['points'])
    #print(team_of_the_day['today']['team'], team_of_the_day['today']['score'])
    #print(live_players_list)

    if player_of_the_day['today']['points'] is  None or player_of_the_day['today']['points'] < 100:
        player_of_the_day_points = player_of_the_day['yesterday']['points']
        player_of_the_day_team = player_of_the_day['yesterday']['team']
        player_of_the_day_name = player_of_the_day['yesterday']['name']

    else:
        player_of_the_day_points = player_of_the_day['today']['points']
        player_of_the_day_team = player_of_the_day['today']['team']
        player_of_the_day_name = player_of_the_day['today']['name']

    if team_of_the_day['today']['score'] is  None  or team_of_the_day['today']['score'] < 100  :
        team_of_the_day_score = team_of_the_day['yesterday']['score']
        team_of_the_day_name = team_of_the_day['yesterday']['team']

    else:
        team_of_the_day_score = team_of_the_day['today']['score']
        team_of_the_day_name = team_of_the_day['today']['team']

    if league == "JAL":
        leaderboard_title = "JAL IPL 2026"
        template_filename = "JAL-IPL2025-Points.html"
    else:
        leaderboard_title = "FPL IPL 2026"
        template_filename = "FPL-IPL2025-Points.html"

    start_date = "22.03.26"
    end_date = "01.06.26"
    todays_date = datetime.now().strftime("%d.%m.%y")
    percent_days_completed = round((datetime.now() - datetime.strptime(start_date, "%d.%m.%y")).days / (datetime.strptime(end_date, "%d.%m.%y") - datetime.strptime(start_date, "%d.%m.%y")).days * 100)


    html_content = """
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <title>{leaderboard_title} Leaderboard</title>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
        <link href="https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;500;600;700&family=Oswald:wght@400;500;600;700&family=Inter:wght@300;400;500;600&display=swap" rel="stylesheet">
       <style>
        /* ── Design Tokens ─────────────────────────────────── */
        :root {{
            --gold:    #FFD700;
            --silver:  #C0C0C0;
            --bronze:  #CD7F32;
            --fire:    #FF6B00;
            --danger:  #FF1744;
            --accent:  #00B4FF;
            --green:   #00E676;
        }}

        /* ── Light Theme ────────────────────────────────────── */
        :root, [data-theme="light"] {{
            --bg:          #F0F2F8;
            --surface:     #FFFFFF;
            --surface2:    #F7F8FC;
            --border:      #E2E6F0;
            --text:        #1A1A2E;
            --text-muted:  #6B7A99;
            --nav-bg:      linear-gradient(135deg, #0F3460 0%, #1A237E 100%);
            --ticker-bg:   linear-gradient(90deg, #0F3460 0%, #162550 50%, #0F3460 100%);
            --card-shadow: 0 4px 24px rgba(26,35,126,0.10);
            --table-head:  #1A237E;
            --table-head-text: #fff;
            --row-hover:   rgba(0,180,255,0.06);
            --rank1-bg:    linear-gradient(90deg, rgba(255,215,0,0.22) 0%, rgba(255,215,0,0.05) 100%);
            --rank2-bg:    linear-gradient(90deg, rgba(192,192,192,0.22) 0%, rgba(192,192,192,0.05) 100%);
            --rank3-bg:    linear-gradient(90deg, rgba(205,127,50,0.22) 0%, rgba(205,127,50,0.05) 100%);
            --danger-bg:   linear-gradient(90deg, rgba(255,23,68,0.12) 0%, rgba(255,23,68,0.03) 100%);
        }}

        /* ── Dark Theme ─────────────────────────────────────── */
        [data-theme="dark"] {{
            --bg:          #0D0D1A;
            --surface:     #161628;
            --surface2:    #1E1E35;
            --border:      #2A2A45;
            --text:        #E8EAFF;
            --text-muted:  #8890BB;
            --nav-bg:      linear-gradient(135deg, #080818 0%, #0F0F2A 100%);
            --ticker-bg:   linear-gradient(90deg, #0A0A20 0%, #14142E 50%, #0A0A20 100%);
            --card-shadow: 0 4px 32px rgba(0,0,0,0.5);
            --table-head:  #0F0F2A;
            --table-head-text: #A0A8D8;
            --row-hover:   rgba(0,180,255,0.07);
            --rank1-bg:    linear-gradient(90deg, rgba(255,215,0,0.18) 0%, rgba(255,215,0,0.04) 100%);
            --rank2-bg:    linear-gradient(90deg, rgba(192,192,192,0.14) 0%, rgba(192,192,192,0.03) 100%);
            --rank3-bg:    linear-gradient(90deg, rgba(205,127,50,0.16) 0%, rgba(205,127,50,0.03) 100%);
            --danger-bg:   linear-gradient(90deg, rgba(255,23,68,0.14) 0%, rgba(255,23,68,0.03) 100%);
        }}

        /* ── Base ───────────────────────────────────────────── */
        *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}

        body {{
            font-family: 'Inter', 'Segoe UI', sans-serif;
            background: var(--bg);
            color: var(--text);
            min-height: 100vh;
            padding-top: 48px;
            transition: background 0.35s, color 0.35s;
        }}

        h1, h2, h3 {{
            font-family: 'Oswald', 'Rajdhani', sans-serif;
            letter-spacing: 0.03em;
        }}

        a {{ color: var(--accent); text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}

        /* ── Ticker ─────────────────────────────────────────── */
        .ticker-wrap {{
            position: fixed;
            top: 0; left: 0;
            width: 100%;
            background: var(--ticker-bg);
            border-bottom: 2px solid var(--fire);
            padding: 8px 0;
            overflow: hidden;
            z-index: 9999;
            white-space: nowrap;
        }}
        .ticker {{
            display: flex;
            width: max-content;
            animation: ticker-scroll 60s linear infinite;
        }}
        .ticker:hover {{ animation-play-state: paused; }}
        .ticker-item {{
            display: inline-flex;
            align-items: center;
            gap: 6px;
            padding: 0 28px;
            font-size: 13px;
            font-weight: 500;
            color: #C8D0F0;
        }}
        .ticker-item::before {{
            content: "▸";
            color: var(--fire);
            font-size: 10px;
        }}
        .ticker-item .player-name {{
            font-weight: 700;
            color: #fff;
            font-family: 'Rajdhani', sans-serif;
            font-size: 14px;
            letter-spacing: 0.04em;
        }}
        .ticker-item .team-tag {{
            color: var(--accent);
            font-size: 11px;
            font-weight: 600;
            background: rgba(0,180,255,0.15);
            padding: 1px 6px;
            border-radius: 4px;
        }}
        .ticker-item .score {{
            color: var(--green);
            font-weight: 700;
            font-size: 14px;
        }}
        @keyframes ticker-scroll {{
            from {{ transform: translateX(0); }}
            to   {{ transform: translateX(-50%); }}
        }}

        /* ── Nav ────────────────────────────────────────────── */
        .site-nav {{
            background: var(--nav-bg);
            padding: 0 24px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            height: 56px;
            box-shadow: 0 2px 16px rgba(0,0,0,0.25);
        }}
        .nav-links {{ display: flex; align-items: center; gap: 4px; }}
        .nav-links a {{
            color: rgba(255,255,255,0.78);
            font-size: 13px;
            font-weight: 500;
            padding: 6px 14px;
            border-radius: 6px;
            transition: background 0.2s, color 0.2s;
            letter-spacing: 0.02em;
        }}
        .nav-links a:hover {{ background: rgba(255,255,255,0.1); color: #fff; text-decoration: none; }}
        .nav-links a.active {{
            background: rgba(255,107,0,0.25);
            color: var(--fire);
            border: 1px solid rgba(255,107,0,0.35);
        }}
        .nav-badge {{
            background: var(--danger);
            color: #fff;
            font-size: 9px;
            font-weight: 700;
            padding: 2px 5px;
            border-radius: 8px;
            margin-left: 4px;
            vertical-align: middle;
            letter-spacing: 0.05em;
        }}
        .theme-btn {{
            background: rgba(255,255,255,0.08);
            border: 1px solid rgba(255,255,255,0.15);
            border-radius: 8px;
            color: rgba(255,255,255,0.8);
            cursor: pointer;
            padding: 7px 11px;
            font-size: 16px;
            transition: background 0.2s;
        }}
        .theme-btn:hover {{ background: rgba(255,255,255,0.15); }}

        /* ── Page Container ─────────────────────────────────── */
        .page-wrap {{
            max-width: 1200px;
            margin: 0 auto;
            padding: 28px 20px 60px;
        }}

        /* ── Page Header ────────────────────────────────────── */
        .page-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            flex-wrap: wrap;
            gap: 16px;
            margin-bottom: 6px;
        }}
        .page-title {{
            font-size: clamp(1.6rem, 4vw, 2.6rem);
            font-weight: 700;
            color: var(--text);
            text-transform: uppercase;
            letter-spacing: 0.06em;
        }}
        .page-title span {{
            background: linear-gradient(90deg, var(--fire), var(--gold));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }}
        .timestamp {{
            font-size: 12px;
            color: var(--text-muted);
            margin-bottom: 28px;
            font-style: italic;
        }}
        .refresh-btn {{
            display: inline-flex;
            align-items: center;
            gap: 8px;
            background: linear-gradient(135deg, var(--fire), #FF9800);
            color: #fff;
            border: none;
            padding: 10px 22px;
            border-radius: 10px;
            font-size: 14px;
            font-weight: 600;
            cursor: pointer;
            box-shadow: 0 4px 16px rgba(255,107,0,0.35);
            transition: transform 0.2s, box-shadow 0.2s;
            letter-spacing: 0.03em;
        }}
        .refresh-btn:hover {{
            transform: translateY(-2px);
            box-shadow: 0 6px 22px rgba(255,107,0,0.5);
        }}
        .refresh-btn i {{ transition: transform 0.5s; }}
        .refresh-btn:hover i {{ transform: rotate(180deg); }}

        /* ── Hero Cards ─────────────────────────────────────── */
        .hero-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
            gap: 20px;
            margin-bottom: 36px;
        }}
        .hero-card {{
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: 18px;
            padding: 24px;
            position: relative;
            overflow: hidden;
            box-shadow: var(--card-shadow);
            transition: transform 0.25s, box-shadow 0.25s;
        }}
        .hero-card:hover {{
            transform: translateY(-4px);
            box-shadow: 0 10px 40px rgba(0,0,0,0.18);
        }}
        .hero-card::before {{
            content: "";
            position: absolute;
            top: 0; left: 0; right: 0;
            height: 4px;
        }}
        .hero-card.team::before  {{ background: linear-gradient(90deg, var(--gold), var(--fire)); }}
        .hero-card.player::before {{ background: linear-gradient(90deg, var(--accent), var(--green)); }}
        .hero-icon {{
            font-size: 28px;
            margin-bottom: 12px;
            display: block;
        }}
        .hero-label {{
            font-size: 11px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.1em;
            color: var(--text-muted);
            margin-bottom: 2px;
        }}
        .hero-value {{
            font-family: 'Oswald', sans-serif;
            font-size: 2rem;
            font-weight: 700;
            color: var(--text);
            line-height: 1.1;
        }}
        .hero-value.score {{
            font-size: 2.6rem;
            background: linear-gradient(135deg, var(--gold), var(--fire));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }}
        .hero-card.player .hero-value.score {{
            background: linear-gradient(135deg, var(--accent), var(--green));
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }}
        .hero-sub {{ font-size: 12px; color: var(--text-muted); margin-top: 4px; }}

        /* ── Section Headers ────────────────────────────────── */
        .section-head {{
            display: flex;
            align-items: center;
            gap: 14px;
            margin: 36px 0 16px;
            flex-wrap: wrap;
        }}
        .section-head h2 {{
            font-size: 1.35rem;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.06em;
            color: var(--text);
        }}
        .season-bar {{
            flex: 1;
            min-width: 120px;
        }}
        .progress {{
            height: 8px;
            background: var(--border);
            border-radius: 99px;
            overflow: hidden;
        }}
        .progress-bar {{
            height: 100%;
            background: linear-gradient(90deg, var(--fire), var(--gold));
            border-radius: 99px;
            transition: width 0.8s ease;
        }}
        .season-pct {{
            font-size: 11px;
            color: var(--text-muted);
            margin-top: 4px;
        }}

        /* ── Tables ─────────────────────────────────────────── */
        .table-container {{ margin: 0 0 28px; overflow-x: auto; }}

        .table {{
            width: 100%;
            border-collapse: separate;
            border-spacing: 0 4px;
            color: var(--text) !important;
            font-size: 14px;
        }}
        .table thead tr {{
            background: transparent;
        }}
        .table th {{
            background: var(--table-head) !important;
            color: var(--table-head-text) !important;
            padding: 12px 16px;
            font-family: 'Rajdhani', sans-serif;
            font-size: 12px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            border: none !important;
        }}
        .table th:first-child {{ border-radius: 8px 0 0 8px; }}
        .table th:last-child  {{ border-radius: 0 8px 8px 0; }}

        .table td {{
            background: var(--surface);
            padding: 13px 16px;
            vertical-align: middle;
            border: none !important;
            border-top: 1px solid var(--border) !important;
            color: var(--text) !important;
            transition: background 0.15s;
        }}
        .table td:first-child {{ border-radius: 8px 0 0 8px; border-left: 3px solid transparent !important; }}
        .table td:last-child  {{ border-radius: 0 8px 8px 0; }}
        .table tbody tr:hover td {{ background: var(--row-hover) !important; }}

        /* ── Leaderboard Row Colors ─────────────────────────── */
        #team-table tbody tr:nth-child(1) td {{
            background: var(--rank1-bg) !important;
            font-weight: 700;
        }}
        #team-table tbody tr:nth-child(1) td:first-child {{
            border-left-color: var(--gold) !important;
        }}
        #team-table tbody tr:nth-child(2) td {{
            background: var(--rank2-bg) !important;
            font-weight: 700;
        }}
        #team-table tbody tr:nth-child(2) td:first-child {{
            border-left-color: var(--silver) !important;
        }}
        #team-table tbody tr:nth-child(3) td {{
            background: var(--rank3-bg) !important;
            font-weight: 700;
        }}
        #team-table tbody tr:nth-child(3) td:first-child {{
            border-left-color: var(--bronze) !important;
        }}
        #team-table tbody tr:nth-child(8) td,
        #team-table tbody tr:nth-child(9) td {{
            background: var(--danger-bg) !important;
        }}
        #team-table tbody tr:nth-child(8) td:first-child,
        #team-table tbody tr:nth-child(9) td:first-child {{
            border-left-color: var(--danger) !important;
        }}

        /* ── Chart & Series Sections ────────────────────────── */
        .chart-container {{ margin: 0 0 28px; }}
        .section-label {{
            font-family: 'Oswald', sans-serif;
            font-size: 1.2rem;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.06em;
            color: var(--text);
            margin: 36px 0 14px;
            padding-bottom: 8px;
            border-bottom: 2px solid var(--border);
        }}

        /* ── View Counter ───────────────────────────────────── */
        .view-counter {{
            position: fixed;
            bottom: 20px; right: 20px;
            background: var(--surface);
            border: 1px solid var(--border);
            color: var(--text-muted);
            padding: 7px 14px;
            border-radius: 20px;
            font-size: 12px;
            display: flex;
            align-items: center;
            gap: 5px;
            z-index: 1000;
            box-shadow: var(--card-shadow);
            animation: fadeUp 0.4s ease;
        }}
        @keyframes fadeUp {{
            from {{ opacity:0; transform: translateY(10px); }}
            to   {{ opacity:1; transform: translateY(0); }}
        }}

        /* ── Responsive ─────────────────────────────────────── */
        @media (max-width: 640px) {{
            .ticker-item {{ padding: 0 16px; font-size: 12px; }}
            .hero-grid {{ grid-template-columns: 1fr; }}
            .page-wrap {{ padding: 20px 14px 50px; }}
        }}

        /* ── Dark mode link color ───────────────────────────── */
        [data-theme="dark"] .table a {{ color: var(--accent); }}
        [data-theme="dark"] .table a:hover {{ color: #7FDDFF; }}

        </style>
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    </head>
    <body>

        <!-- ── Ticker ───────────────────────────────────────── -->
        <div class="ticker-wrap">
            <div class="ticker">{ticker_content}</div>
        </div>

        <!-- ── Nav ──────────────────────────────────────────── -->
        <nav class="site-nav">
            <div class="nav-links">
                <a href="/">Home</a>
                <a href="/live-scoring" class="active">
                    Points Table <span class="nav-badge">LIVE</span>
                </a>
            </div>
            <button class="theme-btn" onclick="toggleTheme()" title="Toggle theme">
                <i class="fas fa-moon" id="theme-icon"></i>
            </button>
        </nav>

        <!-- ── Main Content ─────────────────────────────────── -->
        <div class="page-wrap">

            <!-- Page Header -->
            <div class="page-header">
                <h1 class="page-title">
                    <span>{leaderboard_title}</span> Leaderboard
                </h1>
                <button class="refresh-btn" onclick="window.location.reload()">
                    <i class="fas fa-sync-alt"></i> Refresh
                </button>
            </div>
            <p class="timestamp">{timestamp}</p>

            <!-- Hero Cards -->
            <div class="hero-grid">
                <div class="hero-card team">
                    <span class="hero-icon">🏆</span>
                    <div class="hero-label">Team of the Day</div>
                    <div class="hero-value">{team_of_the_day_name}</div>
                    <div class="hero-value score">{team_of_the_day_score}</div>
                    <div class="hero-sub">pts today</div>
                </div>
                <div class="hero-card player">
                    <span class="hero-icon">⚡</span>
                    <div class="hero-label">Player of the Day</div>
                    <div class="hero-value">{player_of_the_day_name}</div>
                    <div class="hero-sub">{player_of_the_day_team}</div>
                    <div class="hero-value score">{player_of_the_day_points}</div>
                    <div class="hero-sub">pts today</div>
                </div>
            </div>

            <!-- Points Table -->
            <div class="section-head">
                <h2>Points Table</h2>
                <div class="season-bar">
                    <div class="progress">
                        <div class="progress-bar" style="width:{percent_days_completed}%"></div>
                    </div>
                    <div class="season-pct">{percent_days_completed}% of season complete</div>
                </div>
            </div>
            <div id="team-table" class="table-container">{team_table}</div>

            <!-- Today's Scores -->
            {scores_live_heading}
            <div class="table-container">{daily_scores_table}</div>

            <!-- Charts -->
            <div class="chart-container">{team_chart}</div>
            <div class="chart-container">{race_to_finish_chart}</div>

            <!-- Scoreboard & Series Stats -->
            <div class="section-label">Match Stats</div>
            <div class="table-container">{sb_tables}</div>
            <div class="table-container">{series_tables}</div>

            <!-- MVPs -->
            <div class="section-label">MVPs</div>
            <div class="chart-container">{player_chart}</div>
            <div class="chart-container">{role_chart}</div>
            <div class="table-container">{player_table}</div>

        </div><!-- /page-wrap -->

        <script>
            function toggleTheme() {{
                const root = document.documentElement;
                const cur  = root.getAttribute('data-theme') || 'dark';
                const next = cur === 'dark' ? 'light' : 'dark';
                root.setAttribute('data-theme', next);
                document.getElementById('theme-icon').className =
                    next === 'dark' ? 'fas fa-moon' : 'fas fa-sun';
                localStorage.setItem('theme', next);
            }}
            // Restore saved theme
            (function() {{
                const saved = localStorage.getItem('theme') || 'dark';
                document.documentElement.setAttribute('data-theme', saved);
                document.getElementById('theme-icon').className =
                    saved === 'dark' ? 'fas fa-moon' : 'fas fa-sun';
            }})();

            // Page view counter
            async function fetchPageViews() {{
                try {{
                    const res  = await fetch('/page-views/live-scoring');
                    const data = await res.json();
                    const n = parseInt(data.views || 0);
                    if (!isNaN(n)) {{
                        const el = document.createElement('div');
                        el.className = 'view-counter';
                        el.innerHTML = '<i class="fas fa-eye"></i><span>' + n + ' views</span>';
                        document.body.appendChild(el);
                    }}
                }} catch(e) {{}}
            }}
            async function postPageView() {{
                try {{
                    await fetch('/page-views/live-scoring', {{
                        method: 'POST',
                        headers: {{'Content-Type': 'application/json'}}
                    }});
                }} catch(e) {{}}
            }}
            document.addEventListener('DOMContentLoaded', function() {{
                postPageView();
                fetchPageViews();
            }});
        </script>
    </body>
    </html>
    """.format(
        leaderboard_title=leaderboard_title,
        percent_days_completed=percent_days_completed,
        player_chart=player_chart,
        player_of_the_day_name=player_of_the_day_name,
        player_of_the_day_points=player_of_the_day_points,
        player_of_the_day_team=player_of_the_day_team,
        player_table=player_table,
        race_to_finish_chart=race_to_finish_chart,
        role_chart=role_chart,
        sb_tables=sb_tables,
        series_tables=series_tables,
        team_chart=team_chart,
        team_of_the_day_name=team_of_the_day_name,
        team_of_the_day_score=team_of_the_day_score,
        team_table=team_table,
        ticker_content=ticker_content,
        scores_live_heading=scores_live_heading,
        timestamp=timestamp,
        daily_scores_table=daily_scores_table
    )
    
    with open(f"templates/{template_filename}", "w", encoding='utf-8') as f:        
        f.write(html_content)
    
    logging.info(f"HTML report generated successfully for league: {league}")
    
def edit_dataframe_values(df, search_str, replace_str):
    # Replace values in all string columns of the dataframe only when exact match
    for column in df.select_dtypes(include=['object']).columns:
        df[column] = df[column].apply(lambda x: replace_str if x == search_str else x)
    return df

# function to check if player has name_array values and use that for replacement
# for example - if df has player name Jaddu and that exists as name_array value for a player
# then replace Jaddu with player name found in Player model
def replace_player_name(df, Player, league="FPL"):
    # if not found then replace with player name found in Player
    #print(df)
    # first find all player rows where name_array is not null not none not blank
    players_with_aliases = Player.query.filter(
        Player.name_array.isnot(None)
    ).all()      

    if not league:
        league="FPL"      

    for index, row in df.iterrows():
        #print(row)
        #print("replace_player_name using", row[0])
        player_name = row[0]
        player = Player.query.filter_by(name=player_name).first()
        if player is None:
            # check if player name is in name_array of any player
            # first find all player rows where name_array is not null not none not blank
            # check if player name is in name_array of any player

            for player in players_with_aliases:
                if player_name in player.name_array:
                    
                    # check if column name in df is bowler and replace that too
                    if 'Bowler' in df.columns:                                      
                        df.at[index, 'Bowler'] = player.name
                    else:
                        df.at[index, 'Player'] = player.name
                    #print("Player name replaced from %s to %s", player_name, player.name)
                    #insert_log_message(f"Player name replaced from  {player_name} to {player.name}") 
                    player_name = player.name
                    break     

        player = Player.query.filter_by(name=player_name).first()
        if player is None :
            logging.error(f"{league} : Player name alias not found for {player_name}")   
            insert_log_message(f"{league} : Player name alias not found for  {player_name}")                    

    return df



def generate_player_profile_url(player_id):
    """
    Generates a player profile URL using a fixed base URL and player ID
    """
    base_url = "https://m.cricbattle.com/Player-Profile?TournamentId=13385&PlayerId="
    return f"{base_url}{player_id}" 


def create_race_to_finish_chart(all_team_points_df):
     # Create line chart
    fig = px.line(all_team_points_df, 
                x='date', 
                y='Best11Points',
                color='Team Name',
                title='Race to Finish',
                labels={'Best11Points': 'Best11Points', 'date': 'Date'},
                line_shape='linear', markers=True)

    fig.update_layout(xaxis_title='Date', 
                    yaxis_title='Points',
                    legend_title='Team Name')

    return fig.to_html(full_html=False)  


def init_player_id_mapping():
    # Connect to SQLite database
    conn = sqlite3.connect('/mnt/sqlite/cricbattle.db' if os.environ.get("WEBSITE_SITE_NAME") else '/mnt/sqlite/cricbattle.db' if os.environ.get("GOOGLE_CLOUD_PROJECT") else 'instance/cricbattle.db')

    # Query replaced_player table and save as list
    cursor = conn.cursor()
    cursor.execute("SELECT league, old_player_name, new_player_name FROM replaced_player")
    player_id_mapping = cursor.fetchall()

    # Close database connection
    conn.close()

    return player_id_mapping



def merge_player_points(player_team_points_df, league="FPL"):

    if not league:
        league = "FPL"

    try:
        player_id_mapping = init_player_id_mapping()
    except Exception as e:
        logging.warning(f"Could not load player_id_mapping: {e}")
        player_id_mapping = []

    # For each mapping
    for liga, old_player, new_player in player_id_mapping:
        #print(" Merge player function - ", liga, league, old_player, new_player)
        # Get old player row
        if liga == league:            
            old_player_row = player_team_points_df[
            player_team_points_df['Player Name'] == old_player
            ]
            
            # Get new player row 
            new_player_row = player_team_points_df[
                player_team_points_df['Player Name'] == new_player
            ]
            
            # If both players exist
            if not old_player_row.empty and not new_player_row.empty:
                # Sum the points
                total_points = old_player_row.iloc[0]['TotalScore'] + new_player_row.iloc[0]['TotalScore']

                # create duplicate row for new player and update the new row
                # Create new row by copying the data
                merged_player_row = pd.DataFrame([new_player_row.iloc[0]], columns=new_player_row.columns)
                                
                merged_player_row['Player Name'] = new_player + " ⭐"                
                merged_player_row['TotalScore'] = total_points
                merged_player_row['merged_player_flag'] = 'yes'

                #print(merged_player_row)
                #print(player_team_points_df)
                #print(total_points)

                logging.info(f"%s: Dataframe updated for replaced player %s with new player %s", league, old_player, new_player)

                player_team_points_df = pd.concat([player_team_points_df, merged_player_row])

                #print(player_team_points_df)


    return player_team_points_df



def main(Player, PlayerRanking, PlayerRankingPerDay, player_of_the_day, team_of_the_day, league="", live_players_list=pd.DataFrame(), live_player_scores_df=pd.DataFrame(), knockout_players_list=pd.DataFrame()):
    #players_df = read_excel_file("players.xlsx")
    #write code to extract players table from cricbattle.db sqllite database and save as dataframe
    # Connect to SQLite database
    #conn = sqlite3.connect('/mnt/sqlite/cricbattle.db')
    
    # Query players table and save as dataframe
    players_df = pd.DataFrame([{
        'name': p.name,
        'team_name': p.team_name, 
        'role': p.role,
        'ipl_team': p.ipl_team,
        'foreign_player': p.foreign_player,
        'first_match_id': p.first_match_id,
        'selling_price': p.selling_price,
        'category': p.category,
        'point_reduction': p.points_reduction

        } for p in Player.query.all()])
    # Close database connection
    #conn.close()    

    players_df = players_df.rename(columns={'name': 'Player Name'})
    players_df = players_df.rename(columns={'team_name': 'Team Name'})
    players_df = players_df.rename(columns={'role': 'Role'})
    players_df = players_df.rename(columns={'ipl_team': 'IPL Team'})

    #player_rankings_df = read_excel_file("player_rankings.xlsx")
    # create player_rankings_df from PlayerRanking model

    player_rankings_df = pd.DataFrame([{
        'PlayerId': pr.PlayerId,
        'PlayerName': pr.PlayerName,
        'PlayerTypeId': pr.PlayerTypeId,
        'PlayerFormId': pr.PlayerFormId,
        'IsOut': pr.IsOut,
        'IsInjured': pr.IsInjured,
        'Price': pr.Price,
        'RealTeamName': pr.RealTeamName,
        'TotalScore': pr.TotalScore,
        'IsShowTrophy': pr.IsShowTrophy,
        'Rank': pr.Rank,
        'PRank': pr.PRank
        
        } for pr in PlayerRanking.query.all()])
    
    player_rankings_per_day_df = pd.DataFrame([{
        'PlayerId': pr.PlayerId,
        'PlayerName': pr.PlayerName,
        'TotalScore': pr.TotalScore,
        'timestamp': pr.timestamp
        } for pr in PlayerRankingPerDay.query.all()])
    
    all_team_points = pd.DataFrame(columns=['date', 'Team Name', 'Best11Points'])

    #print(player_rankings_per_day_df.head())
    if not players_df.empty and not player_rankings_per_day_df.empty:
        try:

            # keep only max timestamp entries per day in player_rankings_per_day_df
            # Convert timestamp to datetime if not already

            # Initialize global variable
            global player_id_mapping
            player_id_mapping = init_player_id_mapping()

            # Extract date from timestamp 
            player_rankings_per_day_df['date'] = player_rankings_per_day_df['timestamp'].dt.date
            #print(player_rankings_per_day_df)  

            # Sort by timestamp descending and keep first row per date per playername
            
            player_rankings_per_day_df = player_rankings_per_day_df.sort_values('timestamp', ascending=False).groupby(['date', 'PlayerName']).first().reset_index() 
            #print(player_rankings_per_day_df)  

            del player_rankings_per_day_df['timestamp']

            #print(player_rankings_per_day_df)         
            
            # Sort by date
            player_rankings_per_day_df = player_rankings_per_day_df.sort_values('date')

            # Get unique dates
            dates = player_rankings_per_day_df['date'].unique()

            #print(dates)

            # Initialize empty list to store daily team points
            daily_team_points = []

            # Loop through each date
            for date in dates:
                # Get data for current date
                #print("processing for date : ", date)
                daily_df = player_rankings_per_day_df[player_rankings_per_day_df['date'] == date]

                #print(daily_df)
                

                merged_df_perday = pd.merge(players_df, daily_df, left_on="Player Name", right_on="PlayerName")

                #print(merged_df_perday)

                # Add Best 11 Points  
                best_11_data_per_day = calculate_best_11(merged_df_perday, league)  

                #print(best_11_data_per_day)
                team_points_df_per_day = merged_df_perday.groupby('Team Name')['TotalScore'].sum().reset_index()  
                team_points_df_per_day.rename(columns={'TotalScore': 'TotalPoints'}, inplace=True) 

                #print(team_points_df_per_day)

                # Add Best 11 Points  
                best_11_dict_per_day = {team: points for team, points, _ in best_11_data_per_day}  
                team_points_df_per_day['Best11Points'] = team_points_df_per_day['Team Name'].map(best_11_dict_per_day)    

                team_points_df_per_day['date'] = date          
                
                
                daily_team_points.append(team_points_df_per_day)
                #print(daily_team_points)


        except Exception as e:
            logging.error(f"An error occurred during race to finish processing: {str(e)}")
            traceback.print_exception(type(e), e, e.__traceback__)

        # Combine all daily points
        all_team_points = pd.concat(daily_team_points) if daily_team_points else pd.DataFrame(columns=['date', 'Team Name', 'Best11Points'])
        #print(all_team_points)
        #exit()

    
  
    #print(player_rankings_df.head())
    if not players_df.empty and not player_rankings_df.empty:        
        try:
            
            # Merge the dataframes
            merged_df = pd.merge(players_df, player_rankings_df, left_on="Player Name", right_on="PlayerName")

                        
            # Apply point reduction if applicable
            if 'point_reduction' in merged_df.columns:
                merged_df['TotalScore'] = merged_df.apply(lambda row: int(row['TotalScore'] - row['point_reduction']) if pd.notna(row['point_reduction']) else int(row['TotalScore']), axis=1)    


            #print(merged_df[merged_df['Player Name'].str.contains('Ben Dwarshuis', case=False)])    

            # merge replaced players
            player_team_points_df_merged = merge_player_points(merged_df, league)
              
            # Add Best 11 Points  
            #print("Calling new best 11")
            best_11_data = calculate_best_11(player_team_points_df_merged, league)  
            team_points_df = merged_df.groupby('Team Name')['TotalScore'].sum().reset_index()  
            team_points_df.rename(columns={'TotalScore': 'TotalPoints'}, inplace=True)  

             
            #print(team_points_df)
            #exit()
            
              
            # Add Best 11 Points  
            best_11_dict = {team: points for team, points, _ in best_11_data}  
            team_points_df['Best11Points'] = team_points_df['Team Name'].map(best_11_dict)  

            #print(best_11_data)
            # Create list of team and player names from best_11_data
            team_players = []
            for team, _, players in best_11_data:
                for player in players:
                    team_players.append({
                        'Team Name': team,
                        'Player Name': player['Player Name']
                    })

            # Convert to DataFrame
            best_11_df = pd.DataFrame(team_players)        
            #print(best_11_df)                
            
  
            # Sort by Best11Points  
            team_points_df.sort_values(by='Best11Points', ascending=False, inplace=True)
              
            # Second table: Points per player per team, grouped by team using team name from players_df  
            player_team_points_df = player_team_points_df_merged.groupby(['PlayerId', 'Team Name', 'Player Name', 'Role', 'IPL Team'])['TotalScore'].sum().reset_index() 
            player_team_points_df.rename(columns={'TotalScore': 'PlayerPoints'}, inplace=True) 
            #print(player_team_points_df.head())
            # Add player profile URLs to the DataFrame
            player_team_points_df['PlayerId'] = player_team_points_df['PlayerId'].apply(generate_player_profile_url) 
            #print(player_team_points_df.head())   
             
           
            # Get individual series stats
            #df_series = update_series_stats.main()

            # Create SQLite connection
            conn = sqlite3.connect('/mnt/sqlite/cricket_stats.db' if os.environ.get("WEBSITE_SITE_NAME") else '/mnt/sqlite/cricket_stats.db' if os.environ.get("GOOGLE_CLOUD_PROJECT") else 'instance/cricket_stats.db')  


            # Query data from scoreboard tables
            df_series = {}

            # Query batting stats
            try:
                df_series["MOST_RUNS"] = pd.read_sql_query("""
                    SELECT * from cricket_most_runs
                """, conn)
            except:
                df_series["MOST_RUNS"] = pd.DataFrame()

            # Query bowling stats  
            try:
                df_series["MOST_WICKETS"] = pd.read_sql_query("""
                    SELECT * from cricket_most_wickets
                """, conn)
            except:
                df_series["MOST_WICKETS"] = pd.DataFrame()

            # Query bowling stats  
            try:
                df_series["MOST_SIXES"] = pd.read_sql_query("""
                    SELECT * from cricket_most_sixes
                """, conn)
            except:
                df_series["MOST_SIXES"] = pd.DataFrame()
     

            # Print first few extracted tables
            for key, df in df_series.items():
                #print(f"\n=== {key} ===")
                #print(df.head())

                # Merge the dataframes
                # Get the column name in df based on position (assuming the column to merge on is always in position 0)
                #print(df)

                # break if df is empty
                if df.empty:
                    break

                
                merge_column = df.columns[0] if len(df.columns) > 0 else None    
                #print("merge_column :", merge_column)            

                
                merged_df = pd.merge(players_df[['Team Name', 'Player Name']], df, left_on="Player Name", right_on=merge_column, how='right')  
                #print(merged_df.head())
                df_series[key] = merged_df

            # Get individual series stats
            


            # Query data from scoreboard tables
            df_scoreboard = {}

            # Query batting stats
            try:
                df_scoreboard["Bat"] = pd.read_sql_query("SELECT * from cricket_bat", conn)
            except Exception:
                df_scoreboard["Bat"] = pd.DataFrame()

            # Query bowling stats  
            try:
                df_scoreboard["Bowl"] = pd.read_sql_query("SELECT * from cricket_bowl", conn)
            except Exception:
                df_scoreboard["Bowl"] = pd.DataFrame()

            # Query fielding stats
            try:
                df_scoreboard["Field"] = pd.read_sql_query("SELECT * from cricket_field", conn)
            except Exception:
                df_scoreboard["Field"] = pd.DataFrame()

            # Query potm stats
            try:
                df_scoreboard["POTM"] = pd.read_sql_query("SELECT * from cricket_potm", conn)
            except Exception:
                df_scoreboard["POTM"] = pd.DataFrame()

            conn.close()  

            # Print first few extracted tables
            for key, df in df_scoreboard.items():
                #print(f"\n=== {key} ===")
                #print(df.head())

                # check if df has no rows
                if df.empty:
                    continue

                # Merge the dataframes
                # Get the column name in df based on position (assuming the column to merge on is always in position 0)
                merge_column = df.columns[0]  # Get the first column in each DataFrame (e.g., 'Batter', 'Player', 'Bowler')
                #print(merge_column)

                edit_dataframe_values(df, "Kohli", "Virat Kohli")
                edit_dataframe_values(df, "Mitchell Santner (c)", "Mitchell Santner")
                edit_dataframe_values(df, "William ORourke", "William O’Rourke")
                edit_dataframe_values(df, "Salman Agha", "Agha Salman")
                edit_dataframe_values(df, "Shaheen Afridi", "Shaheen Shah Afridi")
                edit_dataframe_values(df, "Latham", "Tom Latham")
                edit_dataframe_values(df, "Tom Latham (wk)", "Tom Latham")
                edit_dataframe_values(df, "Latham (wk)", "Tom Latham")
                edit_dataframe_values(df, "Rahul", "KL Rahul")
                edit_dataframe_values(df, "Rizwan", "Mohammad Rizwan")
                edit_dataframe_values(df, "Shami", "Mohammed Shami")
                edit_dataframe_values(df, "Shanto", "Najmul Hossain Shanto")
                edit_dataframe_values(df, "Shanto (c)", "Najmul Hossain Shanto")
                edit_dataframe_values(df, "Williamson", "Kane Williamson")
                edit_dataframe_values(df, "Azmatullah", "Azmatullah Omarzai")
                edit_dataframe_values(df, "Bavuma", "Temba Bavuma")
                edit_dataframe_values(df, "Temba Bavuma (c)", "Temba Bavuma")
                edit_dataframe_values(df, "Bavuma (c)", "Temba Bavuma")
                edit_dataframe_values(df, "Maharaj", "Keshav Maharaj")
                edit_dataframe_values(df, "Shahidi", "Hashmatullah Shahidi")
                edit_dataframe_values(df, "Rickelton (wk)", "Ryan Rickelton")
                edit_dataframe_values(df, "Rickelton", "Ryan Rickelton")
                edit_dataframe_values(df, "van der Dussen", "Rassie van der Dussen")
                edit_dataframe_values(df, "Markram", "Aiden Markram")
                edit_dataframe_values(df, "Duckett", "Ben Duckett")
                edit_dataframe_values(df, "Josh Inglis (wk)", "Josh Inglis")
                edit_dataframe_values(df, "Labuschagne", "Marnus Labuschagne")
                edit_dataframe_values(df, "Zampa", "Adam Zampa")
                edit_dataframe_values(df, "Maxwell", "Glen Maxwell")
                edit_dataframe_values(df, "Livingstone", "Liam Livingstone")
                edit_dataframe_values(df, "Rabada", "Kagiso Rabada")
                edit_dataframe_values(df, "Rahmat", "Rahmat Shah")
                edit_dataframe_values(df, "Axar", "Axar Patel")
                edit_dataframe_values(df, "Conway", "Devon Conway")
                edit_dataframe_values(df, "Santner (c)", "Mitchell Santner")
                edit_dataframe_values(df, "Santner", "Mitchell Santner")
                edit_dataframe_values(df, "Root", "Joe Root")
                edit_dataframe_values(df, "Dwarshuis", "Ben Dwarshuis")
                edit_dataframe_values(df, "Nabi", "Mohammad Nabi")
                edit_dataframe_values(df, "Gurbaz", "Rahmanullah Gurbaz")
                edit_dataframe_values(df, "Glen Maxwell", "Glenn Maxwell")
                edit_dataframe_values(df, "Gulbadin", "Gulbadin Naib")
                edit_dataframe_values(df, "Klaasen", "Heinrich Klaasen")
                edit_dataframe_values(df, "Mulder", "Wiaan Mulder")
                edit_dataframe_values(df, "Rohit Sharma (c)", "Rohit Sharma")
                edit_dataframe_values(df, "Steven Smith (c)", "Steven Smith")
                edit_dataframe_values(df, "Heinrich Klaasen (wk)", "Heinrich Klaasen")

                #edit_dataframe_values(df, "Philip Salt", "Phil Salt")

                edit_dataframe_values(df, "Ajinkya Rahane (c)", "Ajinkya Rahane")

                edit_dataframe_values(df, "Rasikh Dar Salam", "Rasikh Salam")

                replace_player_name(df, Player, league)

                merged_df = pd.merge(players_df[['Team Name', 'Player Name', 'first_match_id']], df, left_on="Player Name", right_on=merge_column, how='right')  
            
                # Filter out players who joined after the match
                #print(merged_df)

                insert_log_message("Before first match id filter")
                insert_log_message(merged_df)
               

                #check if mattchId col exists in merged_df
                if 'matchId' in merged_df.columns:

                    merged_df = merged_df[~((pd.notna(merged_df['first_match_id'])) & (merged_df['first_match_id'].astype(float) > merged_df['matchId'].astype(float)))]   

                    del merged_df['first_match_id']
                    del merged_df["matchId"]

                    logging.info(f"Removed entries for replaced players for {key}")    
                    insert_log_message(f"Removed entries for replaced players for {key}")

                insert_log_message("After first match id filter")

                insert_log_message(merged_df)     
                
                if "Field" in key:
                    #print(merged_df)
                    # For fielding stats, aggregate by player name first
                    player_catches = merged_df.groupby(['Team Name', 'Player'])['Catches'].sum().reset_index(name='Player Count')                    
                    # keep only one player per team with max Player Count
                    player_catches = player_catches.groupby('Team Name').apply(lambda x: x.nlargest(1, 'Player Count')).reset_index(drop=True)
                    player_catches = player_catches.sort_values('Player Count', ascending=False)
                    player_catches.index = range(1, len(player_catches) + 1)
                    df_scoreboard[key] = player_catches     
                    #print(player_catches)
                
                    
                else:
                    #df_scoreboard[key] = merged_df
                    #print(merged_df)
                    team_counts = merged_df.groupby('Team Name').size().reset_index(name='Player Count')
                    team_counts = team_counts.sort_values('Player Count', ascending=False)
                    team_counts.index = range(1, len(team_counts) + 1)     
                    df_scoreboard[key] = team_counts           
                
                #print(df_scoreboard[key])      


            # add lead_by col to team_points_df which is difference between best11points of current team vs next team
            team_points_df['Lead'] = team_points_df['Best11Points'].diff(-1).fillna(0).abs().astype(int)                                   
            

            # add avg points per day col to team_points_df which is average points per day of current team
            # days is number of distinct days in date col in all_team_points df
            if not all_team_points.empty and 'date' in all_team_points.columns:
                all_team_points_days = all_team_points['Best11Points'].groupby(all_team_points['date']).sum().reset_index()
                days = max(len(all_team_points_days), 1)
            else:
                days = 1
            #print(all_team_points)
            team_points_df['DailyAvg'] = team_points_df['Best11Points'] / days
            #print(team_points_df)
                                     
                
                                            
            # Generate HTML report
            generate_html_report(team_points_df, player_team_points_df, df_series, df_scoreboard, best_11_df, player_of_the_day, team_of_the_day, league, live_players_list, all_team_points, live_player_scores_df)
            
            logging.info("Data transformation and HTML generation complete.")
            
        except Exception as e:
            logging.error(f"An error occurred during data processing: {str(e)}")
            traceback.print_exception(type(e), e, e.__traceback__)
    else:
        logging.error("Data processing aborted due to previous errors.")


def init_player_df(Player, PlayerRanking):
    # Query players table and save as dataframe
    players_df = pd.DataFrame([{
        'name': p.name,
        'team_name': p.team_name, 
        'role': p.role,
        'ipl_team': p.ipl_team,
        'foreign_player': p.foreign_player,
        'first_match_id': p.first_match_id,
        'selling_price': p.selling_price,
        'category': p.category,
        'point_reduction': p.points_reduction

        } for p in Player.query.all()])
    # Close database connection
    #conn.close()    

    players_df = players_df.rename(columns={'name': 'Player Name'})
    players_df = players_df.rename(columns={'team_name': 'Team Name'})
    players_df = players_df.rename(columns={'role': 'Role'})
    players_df = players_df.rename(columns={'ipl_team': 'IPL Team'})

    #player_rankings_df = read_excel_file("player_rankings.xlsx")
    # create player_rankings_df from PlayerRanking model

    player_rankings_df = pd.DataFrame([{
        'PlayerId': pr.PlayerId,
        'PlayerName': pr.PlayerName,
        'PlayerTypeId': pr.PlayerTypeId,
        'PlayerFormId': pr.PlayerFormId,
        'IsOut': pr.IsOut,
        'IsInjured': pr.IsInjured,
        'Price': pr.Price,
        'RealTeamName': pr.RealTeamName,
        'TotalScore': pr.TotalScore,
        'IsShowTrophy': pr.IsShowTrophy,
        'Rank': pr.Rank,
        'PRank': pr.PRank
        
        } for pr in PlayerRanking.query.all()])
    
     # Initialize global variable
    global player_id_mapping
    player_id_mapping = init_player_id_mapping()
    
    return players_df, player_rankings_df


def create_points_table(Player, PlayerRanking):

    players_df, player_rankings_df = init_player_df(Player, PlayerRanking)

    #print(player_rankings_df.head())
    if not players_df.empty and not player_rankings_df.empty:        
        try:
            
            # Merge the dataframes
            merged_df = pd.merge(players_df, player_rankings_df, left_on="Player Name", right_on="PlayerName")

                        
            # Apply point reduction if applicable
            if 'point_reduction' in merged_df.columns:
                merged_df['TotalScore'] = merged_df.apply(lambda row: int(row['TotalScore'] - row['point_reduction']) if pd.notna(row['point_reduction']) else int(row['TotalScore']), axis=1)    

            # Add Best 11 Points  
            player_team_points_df_merged = merge_player_points(merged_df, "FPL")
            best_11_data = calculate_best_11(player_team_points_df_merged, "FPL")  
            team_points_df = merged_df.groupby('Team Name')['TotalScore'].sum().reset_index()  
            team_points_df.rename(columns={'TotalScore': 'TotalPoints'}, inplace=True)  

            # Add Best 11 Points  
            best_11_dict = {team: points for team, points, _ in best_11_data}  
            team_points_df['Best11Points'] = team_points_df['Team Name'].map(best_11_dict)  

            # Sort by Best11Points  
            team_points_df.sort_values(by='Best11Points', ascending=False, inplace=True)
              
            # add lead_by col to team_points_df which is difference between best11points of current team vs next team
            team_points_df['Lead'] = team_points_df['Best11Points'].diff(-1).fillna(0).abs().astype(int)         

            player_team_points_df = merged_df.groupby(['PlayerId', 'Team Name', 'Player Name', 'Role', 'IPL Team'])['TotalScore'].sum().reset_index() 
            player_team_points_df.rename(columns={'TotalScore': 'PlayerPoints'}, inplace=True)    
            player_team_points_df.sort_values(by='PlayerPoints', ascending=False, inplace=True)                       
            
    
            #print(team_points_df)
        except Exception as e:
            logging.error(f"An error occurred during data processing: {str(e)}")
            traceback.print_exception(type(e), e, e.__traceback__)
    else:
        logging.error("Data processing aborted due to previous errors.")

    return team_points_df, player_team_points_df


def get_series_stats(Player, PlayerRanking):
    
    players_df, player_rankings_df = init_player_df(Player, PlayerRanking)

    #print(player_rankings_df.head())
    if not players_df.empty and not player_rankings_df.empty:        
        try:

            # Create SQLite connection
            conn = sqlite3.connect('/mnt/sqlite/cricket_stats.db' if os.environ.get("WEBSITE_SITE_NAME") else '/mnt/sqlite/cricket_stats.db' if os.environ.get("GOOGLE_CLOUD_PROJECT") else 'instance/cricket_stats.db')  


            # Query data from scoreboard tables
            df_series = {}

            # Query batting stats
            try:
                df_series["MOST_RUNS"] = pd.read_sql_query("""
                    SELECT * from cricket_most_runs
                """, conn)
            except:
                df_series["MOST_RUNS"] = pd.DataFrame()

            # Query bowling stats  
            try:
                df_series["MOST_WICKETS"] = pd.read_sql_query("""
                    SELECT * from cricket_most_wickets
                """, conn)
            except:
                df_series["MOST_WICKETS"] = pd.DataFrame()

            # Query bowling stats  
            try:
                df_series["MOST_SIXES"] = pd.read_sql_query("""
                    SELECT * from cricket_most_sixes
                """, conn)
            except:
                df_series["MOST_SIXES"] = pd.DataFrame()


            # Print first few extracted tables
            for key, df in df_series.items():
                #print(f"\n=== {key} ===")
                #print(df.head())

                # Merge the dataframes
                # Get the column name in df based on position (assuming the column to merge on is always in position 0)
                #print(df)

                # break if df is empty
                if df.empty:
                    break

                
                merge_column = df.columns[0] if len(df.columns) > 0 else None    
                #print("merge_column :", merge_column)            

                
                merged_df = pd.merge(players_df[['Team Name', 'Player Name']], df, left_on="Player Name", right_on=merge_column, how='right')  
                #print(merged_df.head())
                df_series[key] = merged_df

        except Exception as e:
            logging.error(f"An error occurred during data processing: {str(e)}")
            traceback.print_exception(type(e), e, e.__traceback__)
    else:
        logging.error("Data processing aborted due to previous errors.")

    # Get individual series stats
    return df_series


def get_scoreboard_stats(Player, PlayerRanking):

    players_df, player_rankings_df = init_player_df(Player, PlayerRanking)
            
    #print(player_rankings_df.head())
    if not players_df.empty and not player_rankings_df.empty:        
        try:

            # Create SQLite connection
            conn = sqlite3.connect('/mnt/sqlite/cricket_stats.db' if os.environ.get("WEBSITE_SITE_NAME") else '/mnt/sqlite/cricket_stats.db' if os.environ.get("GOOGLE_CLOUD_PROJECT") else 'instance/cricket_stats.db')  


            # Query data from scoreboard tables
            df_scoreboard = {}

            # Query batting stats
            try:
                df_scoreboard["Bat"] = pd.read_sql_query("SELECT * from cricket_bat", conn)
            except Exception:
                df_scoreboard["Bat"] = pd.DataFrame()

            # Query bowling stats  
            try:
                df_scoreboard["Bowl"] = pd.read_sql_query("SELECT * from cricket_bowl", conn)
            except Exception:
                df_scoreboard["Bowl"] = pd.DataFrame()

            # Query fielding stats
            try:
                df_scoreboard["Field"] = pd.read_sql_query("SELECT * from cricket_field", conn)
            except Exception:
                df_scoreboard["Field"] = pd.DataFrame()

            try:
                df_scoreboard["Field2"] = pd.read_sql_query("SELECT * from cricket_field", conn)
            except Exception:
                df_scoreboard["Field2"] = pd.DataFrame()

            # Query potm stats
            try:
                df_scoreboard["POTM"] = pd.read_sql_query("SELECT * from cricket_potm", conn)
            except Exception:
                df_scoreboard["POTM"] = pd.DataFrame()

            conn.close()  

            # Print first few extracted tables
            for key, df in df_scoreboard.items():
                #print(f"\n=== {key} ===")
                #print(df.head())

                if df.empty:
                    continue

                # Merge the dataframes
                # Get the column name in df based on position (assuming the column to merge on is always in position 0)
                merge_column = df.columns[0]  # Get the first column in each DataFrame (e.g., 'Batter', 'Player', 'Bowler')
                #print(merge_column)

                # check if df has no rows
                if df.empty:
                    continue


                replace_player_name(df, Player)

                merged_df = pd.merge(players_df[['Team Name', 'Player Name', 'first_match_id']], df, left_on="Player Name", right_on=merge_column, how='right')  
               

                #check if mattchId col exists in merged_df
                if 'matchId' in merged_df.columns:

                    merged_df = merged_df[~((pd.notna(merged_df['first_match_id'])) & (merged_df['first_match_id'].astype(float) > merged_df['matchId'].astype(float)))]   

                    del merged_df['first_match_id']
                    del merged_df["matchId"]

                    logging.info(f"Removed entries for replaced players for {key}")       
                
                if "Field2" in key:
                    # For fielding stats, aggregate by player name first
                    #print("Calcuing field2 now")
                    player_team_catches = merged_df.groupby(['Team Name', 'Player'])['Catches'].sum().reset_index(name='Player Count')
                    #print(player_team_catches.to_string(index=True))                                        
                    #print(player_team_catches)
                    # aggregate by team name
                    team_catches = player_team_catches.groupby('Team Name')['Player Count'].sum().reset_index(name='Player Count')
                    #print(team_catches)
                    team_catches = team_catches.sort_values('Player Count', ascending=False)
                    #print(team_catches)
                    team_catches.index = range(1, len(team_catches) + 1)
                    #print(team_catches)
                    df_scoreboard[key] = team_catches
                    #print(df_scoreboard[key])
                elif "Field" in key:
                    #print(merged_df)
                    # For fielding stats, aggregate by player name first
                    player_catches = merged_df.groupby(['Team Name', 'Player'])['Catches'].sum().reset_index(name='Player Count')                    
                    player_catches = player_catches.sort_values('Player Count', ascending=False)
                    player_catches.index = range(1, len(player_catches) + 1)
                    df_scoreboard[key] = player_catches     
                    #print(player_catches)
                else:
                    #df_scoreboard[key] = merged_df
                    #print(merged_df)
                    team_counts = merged_df.groupby('Team Name').size().reset_index(name='Player Count')
                    team_counts = team_counts.sort_values('Player Count', ascending=False)
                    team_counts.index = range(1, len(team_counts) + 1)     
                    df_scoreboard[key] = team_counts           
                

        except Exception as e:
            logging.error(f"An error occurred during data processing: {str(e)}")
            traceback.print_exception(type(e), e, e.__traceback__)
    else:
        logging.error("Data processing aborted due to previous errors.")

    # Get individual series stats
    return df_scoreboard  





if __name__ == "__main__":
    main()
